from django.shortcuts import render_to_response, render
from django.template import RequestContext

from openpyxl import load_workbook
import openpyxl
from Main.models import ProtectedCountry
from Main.Logic.SearchCountryYear import draw_chart


def get_year_country_from_work_book(wb, year, country):
    ws = wb['ESTIMATES']
    year_column = int(year) - 1950 + 5
    # i = 0
    # for row in ws.rows:
    # if row[0].value == 'Index':
    # for cell in row:
    # if cell.value == year:
    # year_column = i
    # i += 1
    count = -1
    for row in ws.rows:
        if row[2].value == country:
            print("found!")
            print(row[year_column].value)
            count = row[year_column].value
    return count


def get_year_country(request, year, country):
    wb = load_workbook(filename='../../Data/WPP2015_POP_F01_2_TOTAL_POPULATION_MALE.xlsx', read_only=True)
    maleCount = get_year_country_from_work_book(wb, year, country)
    if maleCount == -1:
        maleCount = 'not found!'
    wb = load_workbook(filename='../../Data/WPP2015_POP_F01_3_TOTAL_POPULATION_FEMALE.xlsx', read_only=True)
    femaleCount = get_year_country_from_work_book(wb, year, country)
    if femaleCount == -1:
        femaleCount = 'not found!'
    return render_to_response('get_year_country.html',
                              {'male_count': maleCount,
                               'female_count': femaleCount},
                              context_instance=RequestContext(request))


def update_information(request):
    if request.method == 'POST':
        if request.POST.get('country', None) and request.POST.get('year', None) and \
                request.POST.get('men', None) and request.POST.get('women', None):
            country = request.POST.get('country', None)
            year = request.POST.get('year', None)
            men = request.POST.get('men', None)
            women = request.POST.get('women', None)

            if ProtectedCountry.objects.filter(name_country=country).count() == 0:
                wb = openpyxl.load_workbook('..\..\Data\WPP2015_POP_F01_2_TOTAL_POPULATION_MALE.xlsx')
                message_action = update_information_popularity_on_year(wb['ESTIMATES'], country, year,
                                                                       men) + '-> Men \n'
                wb.save('..\..\Data\WPP2015_POP_F01_2_TOTAL_POPULATION_MALE.xlsx')

                wb = openpyxl.load_workbook('..\..\Data\WPP2015_POP_F01_3_TOTAL_POPULATION_FEMALE.xlsx')
                message_action += update_information_popularity_on_year(wb['ESTIMATES'], country, year,
                                                                        women) + '-> Women'
                wb.save('..\..\Data\WPP2015_POP_F01_3_TOTAL_POPULATION_FEMALE.xlsx')
                return render_to_response('update_information.html', {'message_action': message_action},
                                          context_instance=RequestContext(request))

            else:
                message_action = country + "'s data is lock! you can't update information"
                return render_to_response('update_information.html', {'message_action': message_action},
                                          context_instance=RequestContext(request))
    return render_to_response('update_information.html', {}, context_instance=RequestContext(request))


def update_information_popularity_on_year(ws, name_country, year, num):
    name_city_col = 3
    year_start_col = 6
    year_row = 17

    i = 18
    while True:
        if ws.cell(row=i, column=name_city_col).value == name_country:
            j = 0
            while True:
                if ws.cell(row=year_row, column=year_start_col + j).value == year:
                    ws.cell(row=i, column=year_start_col + j).value = num
                    return 'update data'

                elif not ws.cell(row=i, column=year_start_col + j).value:
                    break

                j += 1

            return 'this year not exist -> year = ' + year

        elif not ws.cell(row=i, column=name_city_col).value:
            break

        i += 1
    return 'this country not exist -> country = ' + name_country


def update_protected_cell_of_country(request):
    if request.method == 'POST':
        message = 'Please enter name of country '

        if request.POST.get('country', None):
            country = request.POST.get('country', None)
            if ProtectedCountry.objects.all().filter(name_country=country).count() > 0:
                message = 'the country exist for action!'
            else:
                p = ProtectedCountry(name_country=country)
                p.save()
                message = 'add country!'
        return render_to_response('protectedCountry.html', {'message': message},
                                  context_instance=RequestContext(request))

    return render_to_response('protectedCountry.html', {}, context_instance=RequestContext(request))


def show_list_population(request):
    if request.method == 'POST' and request.POST.get('year', None):

        year = request.POST.get('year')

        wb = openpyxl.load_workbook('..\..\Data\WPP2015_POP_F01_2_TOTAL_POPULATION_MALE.xlsx')
        m = get_list_population(wb['ESTIMATES'], year)

        wb = openpyxl.load_workbook('..\..\Data\WPP2015_POP_F01_3_TOTAL_POPULATION_FEMALE.xlsx')
        w = get_list_population(wb['ESTIMATES'], year)

        if m and w:
            list_pop = []
            for x in w.keys():
                list_pop.append([x, w[x], m[x], w[x] + m[x]])

            list_pop.sort(key=lambda y: y[3])

            return render_to_response('showListOfPopularity.html', {'year': year, 'list_of_popularity': list_pop},
                                      context_instance=RequestContext(request))

        return render_to_response('showListOfPopularity.html', {'year': year}, context_instance=RequestContext(request))

    return render_to_response('showListOfPopularity.html', {}, context_instance=RequestContext(request))


def get_list_population(ws, year):
    name_city_col = 3
    name_city_row = 29
    year_start_col = 6
    year_row = 17
    list_pop = {}

    i = 0
    while True:
        if ws.cell(row=year_row, column=year_start_col + i).value == year:
            j = 0
            while True:
                if not ws.cell(row=name_city_row + j, column=name_city_col).value:
                    break

                else:
                    print(ws.cell(row=name_city_row + j, column=year_start_col + i).value)
                    list_pop[ws.cell(row=name_city_row + j, column=name_city_col).value] = \
                        int(ws.cell(row=name_city_row + j, column=year_start_col + i).value)

                j += 1

            return list_pop

        elif not ws.cell(row=year_row, column=year_start_col + i).value:
            break

        i += 1

    return None


def show_list_prop(request):
    if request.method == 'POST' and request.POST.get('year', None) and request.POST.get('prop', None):

        year = request.POST.get('year')
        prop = request.POST.get('prop')

        try:
            wb = openpyxl.load_workbook('..\..\Data\WPP2015_POP_F01_2_TOTAL_POPULATION_MALE.xlsx')
            m = get_list_population(wb[prop], year)

            wb = openpyxl.load_workbook('..\..\Data\WPP2015_POP_F01_3_TOTAL_POPULATION_FEMALE.xlsx')
            w = get_list_population(wb[prop], year)

            if m and w:
                list_pop = []
                for x in w.keys():
                    list_pop.append([x, w[x], m[x], w[x] + m[x]])

                list_pop.sort(key=lambda y: y[3])

                return render_to_response('showListOfPopularityWithProp.html',
                                          {'year': year, 'prop': prop, 'list_of_popularity': list_pop},
                                          context_instance=RequestContext(request))

            return render_to_response('showListOfPopularityWithProp.html', {'year': year},
                                      context_instance=RequestContext(request))
        except KeyError:
            print('bad prop!')
            return render_to_response('showListOfPopularityWithProp.html', {'error': 'bad property :|'},
                                      context_instance=RequestContext(request))

    return render_to_response('showListOfPopularityWithProp.html', {}, context_instance=RequestContext(request))


def countryshowchart(request, countryname):
    print(countryname)
    return render_to_response('year_chart.html', {"url": draw_chart(countryname)},
                              context_instance=RequestContext(request))

