__author__ = 'sepideh'


import openpyxl


def update_information_popularity_on_year(ws, year):
    name_city_col = 3
    name_city_row = 29
    year_start_col = 6
    year_row = 17
    list_pop = []

    i = 0
    while True:
        if ws.cell(row=year_row, column=year_start_col+i).value == year:
            j = 0
            while True:
                if not ws.cell(row=name_city_row+j, column=name_city_col).value:
                    break

                else:
                    list_pop.append([ws.cell(row=name_city_row+j, column=name_city_col).value,
                                    int(ws.cell(row=name_city_row+j, column=year_start_col+i).value)])

                j += 1

            list_pop
            return list_pop

        elif not ws.cell(row=year_row, column=year_start_col+i).value:
            break

        i += 1

    return None





wb = openpyxl.load_workbook('..\..\..\Data\WPP2015_POP_F01_2_TOTAL_POPULATION_MALE.xlsx')
ws = wb['ESTIMATES']
print(update_information_popularity_on_year(wb['ESTIMATES'],'kkkk',10,1))



