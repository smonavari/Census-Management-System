__author__ = 'smonavari'

from openpyxl import load_workbook
import plotly.plotly as py
import plotly.graph_objs as go



def get_data_male(name):
    wb = load_workbook(filename="../../Data/WPP2015_POP_F01_2_TOTAL_POPULATION_MALE.xlsx")
    ws = wb['ESTIMATES']
    numbers=list()
    cnt = -1
    found=False
    for row in ws.rows:
        found=False
        cnt=-1
        for cell in row:
            if cell.value:
                cnt += 1
                if name.lower() in str(cell.value).lower():
                    found=True
                    print(cell.value)
                    cnt = 0
                if cnt >= 68 or cnt <= 1:
                    continue
                if found:
                    numbers+=[cell.value]
    return numbers

def get_data_female(name):
    wb = load_workbook(filename="../../Data/WPP2015_POP_F01_3_TOTAL_POPULATION_FEMALE.xlsx")
    ws = wb['ESTIMATES']
    numbers=list()
    cnt = -1
    found=False
    for row in ws.rows:
        found=False
        cnt=-1
        for cell in row:
            if cell.value:
                cnt += 1
                if name.lower() in str(cell.value).lower():
                    found=True
                    print(cell.value)
                    cnt = 0
                if cnt >= 68 or cnt <= 1:
                    continue
                if found:
                    numbers+=[cell.value]
    return numbers


def draw_chart(name):
    country_fe=get_data_female(name)
    print(country_fe)
    country_ma=get_data_male(name)
    print(country_ma)

    trace1 = go.Bar(
        y=list(range(1950, 2016)),
        x=country_fe,
        name='female',
        orientation = 'h',
        marker = dict(
            color = 'rgba(55, 128, 191, 0.6)',
            line = dict(
                color = 'rgba(55, 128, 191, 1.0)',
                width = 1,
            )
        )
    )
    trace2 = go.Bar(
    y=list(range(1950,2016)),
    x=country_ma,
    name='male',
    orientation = 'h',
    marker = dict(
        color = 'rgba(255, 153, 51, 0.6)',
        line = dict(
            color = 'rgba(255, 153, 51, 1.0)',
            width = 1,
        )
    )
    )
    data = [trace1, trace2]
    layout = go.Layout(
        barmode='stack'
    )
    fig = go.Figure(data=data, layout=layout)
    plot_url = py.plot(fig, filename='marker-h-bar')
    return plot_url

